import csv
import os
import gensim
import tweepy
from gensim.utils import simple_preprocess
from spacy.lang.en import stop_words
import sys
CONSUMER_KEY ='1ytH5Xmp4yjdcknJKjr010WCn'
CONSUMER_SECRET ='UBqEskT2EpeDDq1l2CJXyVVHuJ5l7fuMsiUOTw8Qd30HoBeCQc'
ACCESS_TOKEN ='1036001209293975558-fKiwewXhEePmi0nRzyu8a6IFGDFFZN'
ACCESS_TOKEN_SECRET ='Khmmg60weJHnL0sIw7pJf3H2Fe4CchmeZfVGWlTyIqpLD'
auth = tweepy.OAuthHandler(CONSUMER_KEY, CONSUMER_SECRET)
auth.set_access_token(ACCESS_TOKEN, ACCESS_TOKEN_SECRET)
api = tweepy.API(auth)
import csv
import time
import pandas as pd
# open the file in the write mode
class API:
    saveFile = open('APIII.csv', 'w', encoding='UTF8')
    write_outfile = csv.writer(saveFile)
    header = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm']
    write_outfile.writerow(header)
    saveFile.close()
    class TweetListener(tweepy.StreamListener):

        def on_status(self, status):

            if hasattr(status, "retweeted_status"):  # Check if Retweet
                try:
                    status = False
                except AttributeError:
                    print(status.retweeted_status.text)
            else:
                try:
                    print(status.extended_tweet["full_text"])
                    saveFile = open('APIII.csv', 'a', encoding='UTF8')

                    saveFile.write(status.extended_tweet["full_text"])

                    saveFile.close()
                except AttributeError:
                    print(status.text)

    runtime = 30
    listener = TweetListener()
    stream = tweepy.Stream(auth=api.auth, listener=listener)
    languages = ['en']
    input_string = str(sys.argv[1])
    print("Enter elements of a list separated by comma "+input_string)
    user_list = input_string.split(",")
    stream.filter(track=user_list, languages=languages, is_async=True)
    time.sleep(runtime)
    stream.disconnect()
    def joinit(df):
        # REMOVE DUPLICATES
        df['tweet'] = df[['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm']].astype(str).agg('-'.join,
                                                                                                            axis=1)
        df['tweet'] = df['tweet'].replace('-nan', '', regex=True)
        df['tweet'] = df['tweet'].replace('nan', '', regex=True)
        df['tweet'].to_csv('APIII.csv')

    dfa = pd.read_csv('APIII.csv', delimiter=',', encoding='UTF-8', low_memory=False, on_bad_lines='skip')

    joinit(dfa)

from nltk.corpus import stopwords
import re
import string

import pandas as pd
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.stem.wordnet import WordNetLemmatizer
from nltk.tokenize import TweetTokenizer
import numpy as np
import nltk


class Preprocessing:

    def preprocess(df):


        # REMOVE DUPLICATES

        df['Remove_Duplicates'] = df['tweet'].drop_duplicates(keep='first')

        # REMOVES REPEATED LETTERS
        def clean(string):
            if (len(string) == 0):
                return ''
            if (set(string) == set(string[0])):
                return ''
            prev = None
            letters = [l for l in string]
            counter = 1
            new = []
            for l in letters:
                if l == prev:
                    counter += 1
                else:
                    if (counter == 2):
                        new.append(prev)
                    counter = 1
                    new.append(l)
                    prev = l
            return ''.join(new)

        df['removerep'] = df['Remove_Duplicates'].fillna('').map(clean)
        # REMOVE NAN
        df = df.replace('', np.nan)
        df = df.dropna(how="any")

        # REMOVES REPEATED ONLY CHARACTER BY ITSELF
        df["removerepeatedstringwords"] = df['removerep'].str.replace(u"h{2,}", "", regex=True).replace("[?]", "",
                                                                                                        regex=True)

        # CLEANS LINKS, HASHTAGS AND MENTIONS AND NUMBERS
        df['Cleaning_Links_Mentions_Hashtags_Numbers'] = df['removerepeatedstringwords'].replace(r'http\S+', '',
                                                                                                 regex=True).replace(
            r'www\S+', '', regex=True).replace('@[\w]*', '', regex=True).replace('#[\w]*', '', regex=True).replace(
            '\d+',
            '',
            regex=True)
        # REMOVES DOTS
        df['removedotscommas'] = df['Cleaning_Links_Mentions_Hashtags_Numbers'].replace("[.,]+", "", regex=True)

        # REMOVES 3 CHARACTERS AND LESS
        df['removewordswith2char'] = df['removedotscommas'].str.findall('\w{4,}').str.join(' ')

        # REMOVES TEXT EMOJIS
        df['cleantextemoji'] = df['removewordswith2char'].str.replace('[^\w\s/:%.,_-]', '', flags=re.UNICODE,
                                                                      regex=True)

        # REMOVES EMOJIS
        df['Remove_Emoji'] = df['removewordswith2char'].astype(str).apply(
            lambda x: x.encode('ascii', 'ignore').decode('ascii'))

        # LOWERS TEXT
        df['Lower_Mesg'] = df['Remove_Emoji'].str.lower()

        # REMOVES STOP WORDS
        stop = stopwords.words('english')
        df['Remove_Stopword'] = df['Lower_Mesg'].astype(str).apply(
            lambda x: ' '.join([word for word in x.split() if word not in (stopwords.words('english'))]))

        # UNIGRAM TWEETS
        tweetokenizer = TweetTokenizer()
        df['Unigram_Tokenized_Tweets'] = df['Remove_Stopword'].astype(str).apply(tweetokenizer.tokenize)

        # BIGRAMS TWEETS
        # REMOVES PUNCTUATION
        punctuation = string.punctuation
        df['No_Punctuation'] = df['Unigram_Tokenized_Tweets'].apply(
            lambda x: [word for word in x if word not in string.punctuation])

        # LEMMATIZE TWEETS
        lemm = WordNetLemmatizer()
        df['Lemmatize_Tweets'] = df['No_Punctuation'].apply(
            lambda lst: [WordNetLemmatizer().lemmatize(word) for word in lst])
        df['Bigram_Tokenized_Tweets'] = df['Lemmatize_Tweets'].astype(str).apply(
            lambda row: list(nltk.bigrams(row.split(' '))))

        # STEMMS TWEETS
        stemmer = PorterStemmer()
        df['Stem_Tweets'] = df['Lemmatize_Tweets'].apply(lambda x: [PorterStemmer().stem(y) for y in x])
        df['Stem_String'] = df['Stem_Tweets'].astype(str).replace("[']+", "", regex=True)

        # PRINTS COLUMN
        print(df['Stem_Tweets'])
        df['Stem_String'] .to_csv('APIII.csv')

    dfa = pd.read_csv('APIII.csv', delimiter=',', encoding='UTF-8',on_bad_lines='skip')

    preprocess(dfa)


from operator import index
from sqlite3 import Row
import spacy
import pytextrank
from openpyxl import load_workbook
import networkx as nx
import matplotlib.pyplot as plt
import graphviz
import pandas as pd

class Feature_Extraction:
    import pandas as pd
    def Textrank(df):
        df['Stem_Stringg']=df["Stem_String"]
        df.to_excel('list.xlsx', index=None, header=True)

        df = load_workbook(filename='list.xlsx')
        sheet = df.active

        def extraact(text, ro):
            nlp = spacy.load("en_core_web_sm")
            nlp.add_pipe("textrank")
            doc = nlp(text)
            indexx = "B" + str(ro)
            dd = ""
            for phrase in doc._.phrases[:1]:
                leng = phrase.text
                if len(leng) > 20:
                    s = text.index(",")
                    tex = leng[0:s]
                    sheet[indexx].value = tex
                    dd = dd + (tex + " ")
                else:
                    sheet[indexx].value = phrase.text
                    dd = dd + (phrase.text + " ")
            return dd

        def preprocess_document(document, sentence_spliter='.', word_spliter=' ', punct_mark=','):
            document = document.lower().strip()
            for pm in punct_mark:
                document = document.replace(pm, '')
            sentences = [sent for sent in document.split(sentence_spliter) if sent != '']
            document = []
            for sent in sentences:
                words = sent.strip().split(word_spliter)
                document.append(words)
            return document

        def get_entities(document):
            unique_words = []
            for sent in document:
                for word in sent:
                    if word not in unique_words:
                        unique_words.append(word)
            return unique_words

        def get_relations(document):
            bigrams = []
            for sent in document:
                for i in range(len(sent) - 1):
                    pair = [sent[i], sent[i + 1]]
                    if pair not in bigrams:
                        bigrams.append(pair)
            return bigrams

        def build_graph(doc):
            pdoc = preprocess_document(doc)
            nodes = get_entities(pdoc)
            edges = get_relations(pdoc)
            G = nx.Graph()
            G.add_nodes_from(nodes)
            G.add_edges_from(edges)
            return G

        def plot_graph(G):
            plt.figure(figsize=(10, 10))
            pos = nx.spring_layout(G)
            nx.draw(G, pos=pos, node_size=500, with_labels=True)
            edge_labels = nx.get_edge_attributes(G, 'HWW')
            nx.draw_networkx_edge_labels(G, pos, edge_labels)
            plt.title("HW")
            #plt.show()
            plt.savefig(".\static\\"+"textrank.png")

        ro = 1
        dec = ""
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            ro = ro + 1
            text = str(row)
            s1 = text.rindex("[") + 4
            s2 = len(text) - 4
            if text[s2] != '"' and text[s2] != ',':
                s2 = len(text) - 3
                s1 = text.rindex("[") + 2
            textt = text[s1:s2]
            tteexxtt = extraact(textt, ro)
            if ro < 152:
                dec += tteexxtt
        df.save('list.xlsx')
        GG = build_graph(dec)
        plot_graph(GG)
        df = pd.read_excel('list.xlsx')
        df.to_csv('APIII.csv', index=None, header=True)

    dfa = pd.read_csv('APIII.csv',on_bad_lines='skip')
    Textrank(dfa)


import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from sklearn.svm import SVC
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score
from sklearn.preprocessing import LabelEncoder
from textblob import TextBlob


class Sentiment_Analysis:
    
    def svm(df):
        def subject(review):
            return TextBlob(review).sentiment.subjectivity
            # function to calculate polarity

        def word_type(review):
            return TextBlob(review).sentiment.polarity

        # function to review the tweets
        def verify(score):
            if score > 0:
                return 'Positive'
            elif score == 0:
                return 'Neutral'
            else:
                return 'Negative'

        df['Subjectivity'] = df['Stem_Stringg'].apply(subject)
        # print(df['Subjectivity'])
        df['Polarity'] = df['Stem_Stringg'].apply(word_type)
        df['Analysis'] = df['Polarity'].apply(verify)
        df.to_csv('APIII.csv')
        # plt.figure(figsize=(15, 7))
        # plt.subplot(1, 3, 1)
        # plt.title("Support Vector Results on Word Classification")
        tb_counts = df.Analysis.value_counts()
        #plt.pie(tb_counts.values, labels=tb_counts.index, explode=(0, 0, 0.25), autopct='%1.1f%%', shadow=False)
        #  plt.show()

        number = LabelEncoder()
        df['Polarity'] = number.fit_transform(df['Polarity'])
        # print(df)
        dff = df['Analysis'].replace(['Neutral', 'Negative', 'Positive'], [-1, 0, 1])
        # print(dff)
        # Polarity
        features = ["Polarity", "Subjectivity"]

        target = "Polarity"
        # print(df[Subjectivity])
        # print(df[target])
        features_train, features_test, target_train, target_test = train_test_split(df[features], dff, test_size=.2000,random_state=35)
        clf = SVC(kernel='linear')
        clf.fit(features_train, target_train)
        pred = clf.predict(features_test)
        print(pred)
        accuracy = accuracy_score(target_test, pred)

        def make_meshgrid(x, y, h=.02):
            x_min, x_max = x.min() - 1, x.max() + 1
            y_min, y_max = y.min() - 1, y.max() + 1
            xx, yy = np.meshgrid(np.arange(x_min, x_max, h), np.arange(y_min, y_max, h))
            return xx, yy

        def plot_contours(ax, clf, xx, yy, **params):
            Z = clf.predict(np.c_[xx.ravel(), yy.ravel()])
            Z = Z.reshape(xx.shape)
            out = ax.contourf(xx, yy, Z, **params)
            return out

        fig, ax = plt.subplots()
        # title for the plots
        title = ('Decision surface of linear SVC ')
        # Set-up grid for plotting.
        xx, yy = make_meshgrid(features_test['Polarity'], features_test['Subjectivity'])

        plot_contours(ax, clf, xx, yy, cmap=plt.cm.coolwarm, alpha=0.8)
        ax.scatter(features_test['Polarity'], features_test['Subjectivity'], c=target_test)
        ax.legend(['Neutral', 'Negative', 'Positive'])
        ax.set_ylabel('Subjectivity')
        ax.set_xlabel('Polarity')

        plt.plot(features_test['Polarity'], features_test['Subjectivity'])

        ###### u can uncomment these lines inorder to check these two results
        # Classification score
        print(classification_report(target_test, pred))
        # Confusion Matrix of the classification
        print(confusion_matrix(target_test, pred))
        print('Accuracy on the words Classfication: ', accuracy)
        #plt.show()
        plt.savefig(".\static\\"+"svm.png")

    dfa = pd.read_csv('APIII.csv', delimiter=',', encoding='UTF-8',on_bad_lines='skip')

    try:
        dataframe = dfa.loc[:10, :]
    except:
      print("warning: dfa.loc not possible.")
    svm(dfa)
#os.remove("AP.csv")


# Importing modules
import pandas as pd
import os

# Read data into papers
class TopicModeling:
    def LDA(df):

        ######################################################################################################################
        # Import the wordcloud library
        from wordcloud import WordCloud
        import matplotlib.pyplot as plt

        long_string = ','.join(list(df['Stem_Stringg'].values))
        wordcloud = WordCloud(background_color="white", max_words=5000, contour_width=3,
                              contour_color='steelblue').generate(
            long_string)
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.axis("off")
        plt.savefig(".\static\\"+"wordsImage.jpg")
        ######################################################################################################################

        import gensim
        from gensim.utils import simple_preprocess
        import nltk
        from nltk.corpus import stopwords

        nltk.download('stopwords')
        stop_words = stopwords.words('english')
        stop_words.extend(['from', 'subject', 're', 'edu', 'use'])

        def sent_to_words(sentences):
            for sentence in sentences:
                # deacc=True removes punctuations
                yield (gensim.utils.simple_preprocess(str(sentence), deacc=True))

        def remove_stopwords(texts):
            return [[word for word in simple_preprocess(str(doc))
                     if word not in stop_words] for doc in texts]

        data = df.Stem_Stringg.values.tolist()
        data_words = list(sent_to_words(data))
        # remove stop words
        data_words = remove_stopwords(data_words)
        # print(data_words[:1][0][:30])

        ######################################################################################################################
        import gensim.corpora as corpora

        # Create Dictionary
        id2word = corpora.Dictionary(data_words)
        # Create Corpus
        texts = data_words
        # Term Document Frequency
        corpus = [id2word.doc2bow(text) for text in texts]
        # View
        # print(corpus[:1][0][:30])

        ######################################################################################################################

        from pprint import pprint

        # number of topics
        num_topicss = 9
        if __name__ == "__main__":


            lda_model = gensim.models.LdaMulticore(corpus=corpus,
                                                   id2word=id2word,
                                                   num_topics=num_topicss)
            # Print the Keyword in the 10 topics
            pprint(lda_model.print_topics())
            doc_lda = lda_model[corpus]

            ######################################################################################################################

            import pyLDAvis.gensim_models
            import pickle
            import pyLDAvis
            import IPython

            # Visualize the topics
            LDAvis_data_filepath = os.path.join(r'decoment_' + str(num_topicss) + '.docx')

            # # this is a bit time consuming - make the if statement True
            # # if you want to execute visualization prep yourself
            if 1 == 1:
                LDAvis_prepared = pyLDAvis.gensim_models.prepare(lda_model, corpus, id2word)
                with open(LDAvis_data_filepath, 'wb') as f:
                    pickle.dump(LDAvis_prepared, f)
            # load the pre-prepared pyLDAvis data from disk
            with open(LDAvis_data_filepath, 'rb') as f:
                LDAvis_prepared = pickle.load(f)
            pyLDAvis.save_html(LDAvis_prepared, ".\static\\"+r'topic_' + str(num_topicss) + '.html')
            LDAvis_prepared
        # Build LDA model



    dfa = pd.read_csv('APIII.csv', delimiter=',', encoding='UTF-8',on_bad_lines='skip')
    LDA(dfa)
    try:
        1/0
    except ZeroDivisionError:
        print("Script finished until last line.")




